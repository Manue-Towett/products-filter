import os
import re
import configparser
from typing import Optional

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows

from utils import Logger

config = configparser.ConfigParser()

with open("./settings/settings.ini", "r") as file:
    config.read_file(file)

INPUT_PATH = config.get("paths", "input")

OUTPUT_PATH = config.get("paths", "output")

SAVE_IMAGE_FILES = re.search(r"true", config.get("images", "save_files"), re.I)

AVAILABILITY = config.get("minimum values", "offers_0_availability")

MINIMUM_ROI = re.search(r"[-]?\d+[.]?(\d+)?", config.get("minimum values", "ROI"))

MINIMUM_OFFER_COUNT = re.search(r"\d+", config.get("minimum values", "OfferCount"))

MINIMUM_RATING = re.search(r"\d+[.]?(\d+)?", config.get("minimum values", "Rating"))

MINIMUM_REVIEW_COUNT = re.search(r"\d+", config.get("minimum values", "ReviewCount"))

MINIMUM_AMAZON_PRICE = re.search(r"\d+[.]?(\d+)?", config.get("minimum values", "AmazonPrice"))

if not os.path.exists(OUTPUT_PATH):
    os.makedirs(OUTPUT_PATH)

class ProductsFilter:
    """Filters products from an excel file based on given settings"""
    def __init__(self) -> None:
        self.logger =  Logger(__class__.__name__)
        self.logger.info("*****Products Filter Started*****")

        self.green_roi_color = "91bf4d"

        self.blacklisted_words = self.__read_blacklist()

        self.logger.info("%s blacklisted items found." % len(self.blacklisted_words))

        self.files = self.__fetch_input_files()

        self.logger.info("%s files found. Filtering..." % len(self.files))

    @staticmethod
    def __fetch_input_files() -> Optional[list[str]]:
        """Fetches the file paths of the excel files that exist in the input path"""
        return [f"{INPUT_PATH}{name}" for name in os.listdir(INPUT_PATH) if name.endswith("xlsx")]

    @staticmethod
    def __read_blacklist() -> list[str]:
        """Reads the blacklist file to retrieve the blacklisted words"""
        with open("./settings/blacklist.txt", "r") as file:
            return list(filter(None, [line.replace("\n", "").strip() for line in file.readlines()]))
    
    def __read_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """Reads an excel file and returns a dataframe"""
        try:
            self.logger.info("Reading file >> %s" % file_path.split("/")[-1])

            df = pd.read_excel(file_path)

            _ = df["Amazon Price"]

            self.logger.info("%s products found." % len(df))

            return df
        
        except:
            self.logger.info("No products found!")

    def __filter_by_amazon_price(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """Filters the products based on the given minimum amazon price"""
        df_final = df

        if MINIMUM_AMAZON_PRICE and df is not None:
            try:
                self.logger.info("Filtering products by amazon price...")

                min_price = float(MINIMUM_AMAZON_PRICE.group())

                df = df.dropna(subset=["Amazon Price"])

                df_price_str = df.astype({"Amazon Price": "str"})

                df_defined = df_price_str[df_price_str["Amazon Price"].str.contains("[^undefined]", regex=True)]

                df_float_price = df_defined.astype({"Amazon Price": "float"})

                df_final = df_float_price[df_float_price["Amazon Price"] >= min_price]

                self.logger.info("%s products remaining." % len(df_final))
                
            except:
                self.logger.error("Encountered an error filtering by amazon price!")
        
        return self.__filter_by_roi(df_final)

    def __filter_by_roi(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """Filters the products based on the given minimum ROI"""
        df_final = df

        if MINIMUM_ROI and df is not None:
            try:
                self.logger.info("Filtering by minimum ROI...")

                min_roi = float(MINIMUM_ROI.group())

                df = df.dropna(subset=["ROI"])

                df_roi_str = df.astype({"ROI": "str"})

                df_defined = df_roi_str[df_roi_str["ROI"].str.contains("[^undefined]", regex=True)]

                df_defined["roi"] = df["ROI"].astype("str").str.replace("%", "")

                df_float_roi = df_defined.astype({"roi": "float"})

                df_filtered = df_float_roi[df_float_roi["roi"] >= min_roi]

                self.logger.info("%s products remaining." % len(df_filtered))  

                df_final = df_filtered.drop(columns=["roi"])              
            
            except:
                self.logger.error("Encountered an error filtering by ROI!")
        
        return self.__filter_by_rating(df_final)

    def __filter_by_rating(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """Filters the products based on the given minimum rating"""
        df_final = df

        if MINIMUM_RATING and df is not None:
            try:
                self.logger.info("Filtering by minimum rating...")

                min_rating = float(MINIMUM_RATING.group())

                df = df.dropna(subset=["Rating"])

                df_rating_str = df.astype({"Rating": "str"})

                df_defined = df_rating_str[df_rating_str["Rating"].str.contains("[^undefined]", regex=True)]

                df_float_rating = df_defined.astype({"Rating": "float"})

                df_filtered = df_float_rating[df_float_rating["Rating"] >= min_rating]

                self.logger.info("%s products remaining." % len(df_filtered))

                df_final = df_filtered 
            
            except:
                self.logger.error("Encountered an error filtering by rating!")

            return self.__filter_by_review_count(df_final)

    def __filter_by_review_count(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """Filters the products based on the given minimum number of reviews"""
        df_final = df

        if MINIMUM_REVIEW_COUNT and df is not None:
            try:
                self.logger.info("Filtering by minimum number of reviews...")

                min_reviews = int(MINIMUM_REVIEW_COUNT.group())

                df = df.dropna(subset=["ReviewCount"])

                df_reviews_str = df.astype({"ReviewCount": "str"})

                df_defined = df_reviews_str[df_reviews_str["ReviewCount"].str.contains("[^undefined]", regex=True)]

                df_float_reviews = df_defined.astype({"ReviewCount": "float"})

                df_filtered = df_float_reviews[df_float_reviews["ReviewCount"] >= min_reviews]

                self.logger.info("%s products remaining." % len(df_filtered))

                df_final = df_filtered

            except:
                self.logger.error("Encountered an error filtering by number of reviews!")
        
        return self.__filter_by_offer_count(df_final)

    def __filter_by_offer_count(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """Filters the products based on the given minimum number of offers"""
        df_final = df

        if MINIMUM_OFFER_COUNT and df is not None:
            try:
                self.logger.info("Filtering by minimum number of offers...")

                min_offers = int(MINIMUM_OFFER_COUNT.group())

                df = df.dropna(subset=["offerCount"])

                df_offers_str = df.astype({"offerCount": "str"})

                df_defined = df_offers_str[df_offers_str["offerCount"].str.contains("[^undefined]", regex=True)]

                df_float_offers = df_defined.astype({"offerCount": "float"})

                df_filtered = df_float_offers[df_float_offers["offerCount"] >= min_offers]

                self.logger.info("%s products remaining." % len(df_filtered))

                df_final = df_filtered
            
            except:
                self.logger.error("Encountered an error filtering by number of offers!")
        
        return self.__filter_by_availability(df_final)

    def __filter_by_availability(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """Filter products by availability"""
        df_final = df

        if re.search(r"stock", AVAILABILITY) and df is not None:
            self.logger.info("Filtering by availability...")

            availability = re.search(r"[a-zA-Z]", AVAILABILITY).group()

            df = df.dropna(subset=["offers/0/availability"])

            df_available = df[df["offers/0/availability"].str.contains(rf"{availability}", regex=True, flags=re.I)]

            self.logger.info("%s products remaining." % len(df_available))

            df_final = df_available
        
        return df_final
    
    def __filter_out_blacklisted_items(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """Filter out blacklisted items from dataframe"""
        if not len(self.blacklisted_words) or df is None:
            return
        
        self.logger.info("Filtering out blacklisted items...")

        df = df.astype({"Amazon Product Title": "str"})
        
        for item in self.blacklisted_words:
            df = df[~df["Amazon Product Title"].str.contains(rf"{item}", flags=re.I, regex=True)]
        
        self.logger.info("Filtered products: {}".format(len(df)))
        
        return df
    
    def __save_filtered_df(self, df: pd.DataFrame, name: str) -> None:
        """Saves data filtered to an excel file"""
        wb = Workbook()
        ws: Worksheet = wb.active

        ws.append(list(df.columns.values))

        wb_images = Workbook()
        ws_images: Worksheet = wb_images.active

        image_headers = [header for header in df.columns.values  if re.search(r"image", str(header), re.I)]

        image_headers.append("Product is a Match?")

        ws_images.append(image_headers)

        for row in ws.iter_rows(max_col=ws.max_column):
            for cell in row:
                cell.font = Font(bold=True)
        
        for row in ws_images.iter_rows(max_col=ws_images.max_column):
            for cell in row:
                cell.font = Font(bold=True)
        
        self.logger.info("Retrieving images from excel...")

        rows = dataframe_to_rows(df, index=False, header=False)

        for row in rows:
            product_images = []

            for cell in row:
                if str(cell).startswith("=IM"):
                    product_images.append(cell)
            
            ws_images.append(product_images)

            ws.append(row)
        
        for row in ws.iter_rows(max_col=ws.max_column):
            for cell in row:
                if str(cell.internal_value).startswith("=HY"):
                    cell.style = "Hyperlink"

                    continue
                
                if str(cell.internal_value).endswith("%"):
                    cell.fill = PatternFill(fill_type="solid",
                                            end_color=self.green_roi_color,
                                            start_color=self.green_roi_color)
        
        self.logger.info("Saving data to excel...")

        wb.save(f'{OUTPUT_PATH}{name.replace(".xlsx", "_filtered.xlsx")}')

        self.logger.info("Filtered data saved to: {}".format(name.replace(".xlsx", "_filtered.xlsx")))

        if SAVE_IMAGE_FILES:
            self.logger.info("Saving images...")

            wb_images.save(f'{OUTPUT_PATH}{name.replace(".xlsx", "_images.xlsx")}')

            self.logger.info("Images saved to: {}".format(name.replace(".xlsx", "_images.xlsx")))

    def run(self) -> None:
        """Entry point to the filter"""
        for file in self.files:
            df = self.__read_file(file_path=file)

            columns = df.columns.values

            wb = load_workbook(file, rich_text=True)

            ws = wb[wb.sheetnames[0]]

            rows = []

            for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
                rows.append([cell.value for cell in row])
            
            unfiltered_df = pd.DataFrame(rows, columns=columns)

            try:
                unfiltered_df.loc[:, "ROI"] = unfiltered_df["ROI"].apply(
                    lambda x: '{:.2%}'.format(x)
                )
            except:pass

            filtered_df = self.__filter_by_amazon_price(unfiltered_df)

            blacklist_filtered_df = self.__filter_out_blacklisted_items(filtered_df)

            if blacklist_filtered_df is None:
                continue
            
            if len(blacklist_filtered_df):
                self.__save_filtered_df(blacklist_filtered_df, file.split("/")[-1])

app = ProductsFilter()
app.run()